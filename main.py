from openpyxl import*
import re
ost = load_workbook('1.xlsx')
b = ost['TDSheet']
a = ost.create_sheet(title="бирки")
sb = []
pr = []
ras = []
for i in range(1,b.max_row + 1):
	q =  b['D' + str(i)].value if b['D' + str(i)].value else 'dfaf' 
	qq = b['K' + str(i)].value
	if re.search('Холодильник хранения туш', q ) and qq == None:
		pr.append(b['B' + str(i)].value)
	elif re.search('сан', q ) and qq == None:
		sb.append(b['B' + str(i)].value)	
	elif qq != None and re.search('Холодильник хранения туш',qq):
		ras.append(b['B' + str(i)].value)
ostn = [i for i in pr if i not in ras]
ostbr = [i for i in sb if i not in ras]
ostn.append('свин')
ostbr.append('Санбрак еще в санкамере?')
for i in list(map(str,ostn)):
	a.append([i])

for i in list(map(str,ostbr)):
	a.append([i])
ost.save('1.xlsx')	